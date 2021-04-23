import os,sys
#shutil是高级处理文件文件夹以及压缩包处理模块
import shutil
#openpyxl是用来读写Excel文件的库
#安装openpyxl之前还需要安装jdcal和et_xmlfile这两个库，才能安装成功
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font,Alignment
from openpyxl.styles.colors import RED,GREEN,DARKYELLOW
from getsqldata.getsqldata import GetsqlData
file = os.path.dirname(__file__)
TARGET_FILE = os.path.dirname(file)+'/report/CskuidAndCspuid.xlsx'
ErrorOrder_FILE = os.path.dirname(file)+'/report/ErrorOrder.xlsx'
MissingCskuid_FILE = os.path.dirname(file)+'/report/MissingCskuid.xlsx'

class WriteExcel():
    """文件写入数据"""
    def write_data(self):
        """
        将查询的商品数据导出到excel中
        :return:
        """
        colx = ['cskuid', 'cspuid']
        if os.path.exists(TARGET_FILE):
            # 文件存在，则删除重新创建一个excel文件
            os.remove(TARGET_FILE)
            # load_workbook接受路径返回一个excel对象，Workbook()创建一个新的excel文件
            self.wb = Workbook()
            # active创建一个sheet工作铺
            self.ws = self.wb.active
            for col in range(len(colx)):
                self.ws.cell(1, col + 1).value = colx[col]
        else:
            self.wb = self.wb = Workbook()
            # active创建一个sheet工作铺
            self.ws = self.wb.active
            for col in range(len(colx)):
                self.ws.cell(1, col + 1).value = colx[col]
        #data为查询的sql元组数据
        getsqldata = GetsqlData()
        data = getsqldata.get_goods()
        for d in data:
            self.ws.append(d)
        self.wb.save(TARGET_FILE)

    def errorOrder_data(self,errorOrder):
        """
         将检测后错误的订单数据导出到excel中
         :return:
          """
        colx = ['orderid','ordercspuid','goodscspuid']
        if os.path.exists(ErrorOrder_FILE):
            # 文件存在，则删除重新创建一个excel文件
            os.remove(ErrorOrder_FILE)
            # load_workbook接受路径返回一个excel对象，Workbook()创建一个新的excel文件
            self.wb = Workbook()
            # active创建一个sheet工作铺
            self.ws = self.wb.active
            for col in range(len(colx)):
                self.ws.cell(1, col + 1).value = colx[col]
        else:
            self.wb = self.wb = Workbook()
            # active创建一个sheet工作铺
            self.ws = self.wb.active
            for col in range(len(colx)):
                self.ws.cell(1, col + 1).value = colx[col]
        for d in errorOrder:
            self.ws.append(d)
        self.wb.save(ErrorOrder_FILE)

    def missingCskuid_data(self, errorOrder):
        """
         将检测后缺失cskuid的订单和cskuid导出到excel中
         :return:
          """
        colx = ['orderid', 'ordercskuid']
        if os.path.exists(MissingCskuid_FILE):
            # 文件存在，则删除重新创建一个excel文件
            os.remove(MissingCskuid_FILE)
            # load_workbook接受路径返回一个excel对象，Workbook()创建一个新的excel文件
            self.wb = Workbook()
            # active创建一个sheet工作铺
            self.ws = self.wb.active
            for col in range(len(colx)):
                self.ws.cell(1, col + 1).value = colx[col]
        else:
            self.wb = self.wb = Workbook()
            # active创建一个sheet工作铺
            self.ws = self.wb.active
            for col in range(len(colx)):
                self.ws.cell(1, col + 1).value = colx[col]
        for d in errorOrder:
            self.ws.append(d)
        self.wb.save(MissingCskuid_FILE)
'''
if __name__ == '__main__':
    WriteExcel().write_data()
'''